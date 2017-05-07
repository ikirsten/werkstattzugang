#Script liest die Daten aus der AMS-Excel Tabelle und schreibt sie in die Datenbank. Alle Daten aus der Datenbank werden zuvor geloescht.

#import
import config #Konfigurationsdatei

from openpyxl import load_workbook #Modul um Excel Dateien zu lesen
import MySQLdb #Modul fuer die Datenbankverbindung

#Tabelle Oeffnen
wb = load_workbook(filename = config.amsdaten, data_only=True)
#Arbeitsblatt auswaehlen
ab = wb['AMS-Daten']



#Datenbank Connection Module
db = MySQLdb.connect(	host=config.database['host'],
			user=config.database['username'],
			passwd=config.database['password'],
			db=config.database['databasename'])

#Datenbank Cursor
c=db.cursor()

a= c.execute("""SELECT * FROM ams_mitglieder""")

#Nur ausfueren, wenn Datensaetze groeser oder gleich der Mitgliedstabelle ist
if a <= ab.max_row+1:
	#####ALLE Daten aus Tabelle loeschen
	c.execute("""DELETE FROM ams_mitglieder""")
	
	for i in xrange(2,ab.max_row+1):
		c.execute("""
		INSERT INTO ams_mitglieder (
		ams_nr, vorname, nachname, ams_mitglied, strasse, land, plz, ort, tel, mobil, email)
		Values (
		%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s); """,
		(int(ab.cell(row=i, column=1).value), ab.cell(row=i, column=2).value, ab.cell(row=i, column=3).value, ab.cell(row=i, column=4).value, ab.cell(row=i, column=10).value, ab.cell(row=i, column=11).value, ab.cell(row=i, column=12).value, ab.cell(row=i, column=13).value, ab.cell(row=i, column=14).value, ab.cell(row=i, column=15).value, ab.cell(row=i, column=16).value))
		print(ab.cell(row=i, column=1).value)
		db.commit()
	
	c.execute(""" INSERT INTO werkelog (
		typ, meldung)
		Values (
		'TUERSYSTEM', 'AMS Datenbank aktualisiert') """)
	db.commit()
	
	
db.close()
